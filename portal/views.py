from datetime import date
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView, LogoutView
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .forms import (
    DepartmentForm,
    PortalPasswordForm,
    PortalUserForm,
    ProfileForm,
    StyledAuthForm,
    SupplierForm,
    SupplyRequestForm,
    WarehouseForm,
)
from .models import (
    Alert,
    DailyVolume,
    Department,
    RequestFile,
    Shipment,
    Supplier,
    SupplyRequest,
    UserProfile,
    Warehouse,
)
from .roles import (
    ROLE_GUIDE,
    ROLE_LABELS,
    apply_role_flags,
    deny_unless,
    require_perm,
    role_label,
)


class PortalLoginView(LoginView):
    template_name = "auth/login.html"
    authentication_form = StyledAuthForm
    redirect_authenticated_user = True


class PortalLogoutView(LogoutView):
    next_page = "login"


def filtered_requests(request):
    qs = (
        SupplyRequest.objects.select_related("supplier", "warehouse")
        .prefetch_related("files")
        .order_by("-created_at", "number")
    )
    status = request.GET.get("status") or ""
    supplier_id = request.GET.get("supplier") or ""
    warehouse_id = request.GET.get("warehouse") or ""
    q = request.GET.get("q") or ""
    if status:
        qs = qs.filter(status=status)
    if supplier_id:
        qs = qs.filter(supplier_id=supplier_id)
    if warehouse_id:
        qs = qs.filter(warehouse_id=warehouse_id)
    if q:
        qs = qs.filter(
            Q(number__icontains=q)
            | Q(supplier__name__icontains=q)
            | Q(supplier__code__icontains=q)
            | Q(warehouse__name__icontains=q)
            | Q(warehouse__city__icontains=q)
        ).distinct()
    return qs, status, supplier_id, warehouse_id, q


PAGE_SIZE = 10


def paginate_qs(request, qs, per_page=PAGE_SIZE):
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page"))
    params = request.GET.copy()
    params.pop("page", None)
    querystring = params.urlencode()
    return page_obj, querystring


def request_file_names(obj):
    names = [item.name for item in obj.files.all()]
    if names:
        return "، ".join(names)
    if obj.attachment:
        return obj.attachment_name
    return ""


def request_route_label(obj):
    if obj.sent_to_supplier_at and obj.sent_to_warehouse_at:
        return "أُرسل للمورد والمستودع"
    return obj.get_status_display()


def build_supply_alerts():
    alerts = []
    qs = SupplyRequest.objects.select_related("supplier", "warehouse")

    for obj in qs.filter(status=SupplyRequest.Status.DELAYED)[:2]:
        warehouse = obj.warehouse.name if obj.warehouse else "المستودع"
        alerts.append(
            {
                "level": "critical",
                "icon": "alert",
                "title": f"تأخر أمر التوريد {obj.number}",
                "detail": f"{obj.supplier.name} — {warehouse} · بانتظار المتابعة",
                "url": reverse("request_detail", args=[obj.pk]),
            }
        )

    awaiting = qs.filter(status=SupplyRequest.Status.SENT).order_by("-sent_to_warehouse_at")[:2]
    for obj in awaiting:
        warehouse = obj.warehouse.name if obj.warehouse else "المستودع"
        alerts.append(
            {
                "level": "info",
                "icon": "docs",
                "title": f"بانتظار تأكيد الاستلام {obj.number}",
                "detail": f"أُرسل إلى {obj.supplier.name} و{warehouse}",
                "url": reverse("request_detail", args=[obj.pk]),
            }
        )

    received = qs.filter(status=SupplyRequest.Status.RECEIVED).order_by("-id")[:1]
    for obj in received:
        warehouse = obj.warehouse.name if obj.warehouse else "المستودع"
        alerts.append(
            {
                "level": "success",
                "icon": "check",
                "title": f"تم استلام أمر التوريد {obj.number}",
                "detail": f"{warehouse} أكّد استلام ملفات التوريد",
                "url": reverse("request_detail", args=[obj.pk]),
            }
        )

    return alerts[:4]


def simple_excel(title, headers, rows, widths, filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.sheet_view.rightToLeft = True
    header_font = Font(name="Calibri", bold=True, color="1E40AF", size=11)
    header_fill = PatternFill("solid", fgColor="E8F1FF")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="right", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    zebra = PatternFill("solid", fgColor="F8FAFC")
    sheet.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = sheet.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin
    for index, row in enumerate(rows, 1):
        sheet.append(row)
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(index + 1, col)
            cell.alignment = cell_align
            cell.border = thin
            cell.font = Font(name="Calibri", size=11)
            if index % 2 == 0:
                cell.fill = zebra
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 28
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_perm("dashboard")
def dashboard(request):
    request.active_nav = "dashboard"
    activities = (
        SupplyRequest.objects.select_related("supplier", "warehouse")
        .prefetch_related("files")
        .order_by("-created_at", "number")[:10]
    )
    volumes = list(DailyVolume.objects.all())
    max_volume = max((item.volume for item in volumes), default=1)
    for item in volumes:
        item.bar_height = int((item.volume / max_volume) * 100)
    kpis = {
        "active_orders": SupplyRequest.objects.exclude(status=SupplyRequest.Status.RECEIVED).count(),
        "occupancy": int(
            Warehouse.objects.aggregate(avg=Sum("capacity_percent"))["avg"] or 0
        )
        // max(Warehouse.objects.count(), 1),
        "pending": SupplyRequest.objects.filter(status=SupplyRequest.Status.PENDING).count(),
        "scheduled": Shipment.objects.exclude(status=Shipment.Status.UNSCHEDULED).count(),
    }
    return render(
        request,
        "dashboard.html",
        {
            "kpis": kpis,
            "volumes": volumes,
            "alerts": build_supply_alerts(),
            "activities": activities,
            "today": timezone.localdate(),
        },
    )


@login_required
@require_perm("requests.view")
def request_list(request):
    request.active_nav = "requests"
    qs, status, supplier_id, warehouse_id, q = filtered_requests(request)

    form = SupplyRequestForm()
    if request.method == "POST":
        blocked = deny_unless(request, "requests.create")
        if blocked:
            return blocked
        form = SupplyRequestForm(request.POST, request.FILES)
        if form.is_valid():
            now = timezone.now()
            obj = form.save(commit=False)
            last = SupplyRequest.objects.order_by("-id").first()
            next_num = (last.id + 1) if last else 1
            obj.number = f"PO-{now.year}-{next_num:03d}"
            obj.created_by = request.user
            obj.placed_at = now
            obj.shipped_at = now
            obj.status = SupplyRequest.Status.SENT
            obj.sent_to_supplier_at = now
            obj.sent_to_warehouse_at = now
            obj.save()
            for uploaded in form.cleaned_data["uploaded_files"]:
                RequestFile.objects.create(request=obj, file=uploaded)
            first = obj.files.first()
            if first:
                obj.attachment = first.file
                obj.save(update_fields=["attachment"])
            messages.success(
                request,
                f"أُرسل طلب التوريد {obj.number} إلى المورد والمستودع",
            )
            return redirect("request_detail", pk=obj.pk)

    page_obj, export_query = paginate_qs(request, qs)
    return render(
        request,
        "requests/list.html",
        {
            "requests_qs": page_obj,
            "page_obj": page_obj,
            "suppliers": Supplier.objects.all(),
            "warehouses": Warehouse.objects.all(),
            "status_choices": SupplyRequest.Status.choices,
            "selected_status": status,
            "selected_supplier": supplier_id,
            "selected_warehouse": warehouse_id,
            "query": q,
            "form": form,
            "total_count": page_obj.paginator.count,
            "export_query": export_query,
        },
    )


@login_required
@require_perm("requests.export")
def export_requests(request):
    qs, _, _, _, _ = filtered_requests(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "طلبات التوريد"
    sheet.sheet_view.rightToLeft = True

    headers = [
        "م",
        "رقم الطلب",
        "المصدر",
        "رقم المورد",
        "اسم المورد",
        "المستودع",
        "المدينة",
        "الملفات",
        "الأولوية",
        "الحالة",
        "التوجيه",
        "تاريخ الإرسال للمورد",
        "تاريخ الإرسال للمستودع",
        "تاريخ الإنشاء",
    ]
    header_font = Font(name="Calibri", bold=True, color="1E40AF", size=11)
    header_fill = PatternFill("solid", fgColor="E8F1FF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    zebra = PatternFill("solid", fgColor="F8FAFC")
    local_tz = timezone.get_current_timezone()

    def fmt_dt(value):
        if not value:
            return ""
        if timezone.is_aware(value):
            value = timezone.localtime(value, local_tz)
        return value.strftime("%Y-%m-%d %H:%M")

    sheet.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = sheet.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    for index, obj in enumerate(qs, 1):
        row = [
            index,
            obj.number,
            "دائرة المشتريات",
            obj.supplier.code if obj.supplier else "",
            obj.supplier.name if obj.supplier else "",
            obj.warehouse.name if obj.warehouse else "",
            obj.warehouse.city if obj.warehouse else "",
            request_file_names(obj),
            obj.get_priority_display(),
            obj.get_status_display(),
            request_route_label(obj),
            fmt_dt(obj.sent_to_supplier_at),
            fmt_dt(obj.sent_to_warehouse_at),
            fmt_dt(obj.created_at),
        ]
        sheet.append(row)
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(index + 1, col)
            cell.alignment = cell_align
            cell.border = thin
            cell.font = Font(name="Calibri", size=11)
            if index % 2 == 0:
                cell.fill = zebra

    widths = [6, 18, 18, 16, 28, 16, 12, 32, 12, 16, 24, 22, 22, 20]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 28

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"talabat-tawreed-{timezone.localdate().isoformat()}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_perm("requests.view")
def request_detail(request, pk):
    request.active_nav = "requests"
    obj = get_object_or_404(
        SupplyRequest.objects.select_related("supplier", "warehouse", "created_by").prefetch_related(
            "files", "items__product"
        ),
        pk=pk,
    )
    return render(request, "requests/detail.html", {"obj": obj})


@login_required
@require_POST
@require_perm("requests.receive")
def mark_received(request, pk):
    obj = get_object_or_404(SupplyRequest, pk=pk)
    obj.status = SupplyRequest.Status.RECEIVED
    obj.save(update_fields=["status"])
    messages.success(request, f"تم تعليم الطلب {obj.number} كمستلم")
    return redirect("request_detail", pk=pk)


@login_required
@require_POST
@require_perm("requests.issue")
def report_issue(request, pk):
    obj = get_object_or_404(SupplyRequest, pk=pk)
    obj.status = SupplyRequest.Status.DELAYED
    obj.save(update_fields=["status"])
    Alert.objects.create(
        level=Alert.Level.CRITICAL,
        title=f"بلاغ تشغيلي على أمر التوريد {obj.number}",
        detail=f"{obj.supplier.name} — أُحيل للمورد والمستودع",
    )
    messages.warning(request, f"تم تسجيل بلاغ على أمر التوريد {obj.number}")
    return redirect("request_detail", pk=pk)


@login_required
@require_perm("schedule.view")
def schedule(request):
    request.active_nav = "schedule"
    target_date = date(2023, 10, 24)
    warehouses = list(Warehouse.objects.all())
    hours = [8, 9, 10, 11]
    scheduled = {
        f"{item.warehouse_id}-{item.scheduled_hour}": item
        for item in Shipment.objects.select_related("warehouse").exclude(
            status=Shipment.Status.UNSCHEDULED
        )
        if item.warehouse_id and item.scheduled_hour
    }
    rows = []
    for hour in hours:
        rows.append(
            {
                "hour": hour,
                "cells": [
                    {
                        "warehouse": warehouse,
                        "shipment": scheduled.get(f"{warehouse.id}-{hour}"),
                    }
                    for warehouse in warehouses
                ],
            }
        )

    return render(
        request,
        "schedule.html",
        {
            "warehouses": warehouses,
            "rows": rows,
            "unscheduled": Shipment.objects.filter(status=Shipment.Status.UNSCHEDULED),
            "target_date": target_date,
            "view": request.GET.get("view", "day"),
        },
    )


@login_required
@require_POST
@require_perm("schedule.assign")
def assign_shipment(request):
    shipment = get_object_or_404(Shipment, pk=request.POST.get("shipment_id"))
    warehouse = get_object_or_404(Warehouse, pk=request.POST.get("warehouse_id"))
    hour = int(request.POST.get("hour"))
    shipment.warehouse = warehouse
    shipment.scheduled_hour = hour
    shipment.scheduled_date = date(2023, 10, 24)
    shipment.status = Shipment.Status.PENDING
    shipment.save()
    if request.headers.get("HX-Request") or request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"ok": True})
    messages.success(request, f"تم جدولة {shipment.number} في {warehouse.name} الساعة {hour:02d}:00")
    return redirect("schedule")


@login_required
@require_perm("settings")
def settings_view(request):
    request.active_nav = "settings"
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    tab = request.GET.get("tab") or "general"
    profile_form = ProfileForm(user=request.user, instance=profile)
    password_form = PortalPasswordForm(user=request.user)

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "profile":
            profile_form = ProfileForm(request.POST, user=request.user, instance=profile)
            if profile_form.is_valid():
                profile_form.save(user=request.user)
                messages.success(request, "تم حفظ التغييرات")
                return redirect("settings")
            tab = "general"
        elif action == "password":
            password_form = PortalPasswordForm(user=request.user, data=request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "تم تحديث كلمة المرور")
                return redirect("settings")
            tab = "security"

    return render(
        request,
        "settings.html",
        {
            "profile": profile,
            "profile_form": profile_form,
            "password_form": password_form,
            "tab": tab,
        },
    )


@login_required
@require_perm("departments.view")
def department_list(request):
    request.active_nav = "setup"
    request.setup_tab = "departments"
    q = request.GET.get("q") or ""
    qs = Department.objects.all().order_by("code")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    form = DepartmentForm()
    if request.method == "POST":
        blocked = deny_unless(request, "departments.manage")
        if blocked:
            return blocked
        form = DepartmentForm(request.POST)
        if form.is_valid():
            item = form.save()
            messages.success(request, f"تمت إضافة القسم {item.name}")
            return redirect("departments")
    page_obj, export_query = paginate_qs(request, qs)
    return render(
        request,
        "setup/departments.html",
        {
            "items": page_obj,
            "page_obj": page_obj,
            "form": form,
            "query": q,
            "export_query": export_query,
        },
    )


@login_required
@require_perm("departments.view")
def export_departments(request):
    q = request.GET.get("q") or ""
    qs = Department.objects.all().order_by("code")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    rows = [[i, item.code, item.name] for i, item in enumerate(qs, 1)]
    return simple_excel(
        "الأقسام",
        ["م", "رقم القسم", "اسم القسم"],
        rows,
        [6, 18, 28],
        f"departments-{timezone.localdate().isoformat()}.xlsx",
    )


@login_required
@require_perm("departments.manage")
def department_edit(request, pk):
    request.active_nav = "setup"
    request.setup_tab = "departments"
    item = get_object_or_404(Department, pk=pk)
    form = DepartmentForm(instance=item)
    if request.method == "POST":
        form = DepartmentForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f"تم تعديل القسم {item.name}")
            return redirect("departments")
    return render(request, "setup/department_edit.html", {"form": form, "item": item})


@login_required
@require_POST
@require_perm("departments.manage")
def department_delete(request, pk):
    item = get_object_or_404(Department, pk=pk)
    name = item.name
    item.delete()
    messages.success(request, f"تم حذف القسم {name}")
    return redirect("departments")


@login_required
@require_perm("warehouses.view")
def warehouse_list(request):
    request.active_nav = "setup"
    request.setup_tab = "warehouses"
    q = request.GET.get("q") or ""
    qs = Warehouse.objects.all().order_by("code")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    form = WarehouseForm()
    if request.method == "POST":
        blocked = deny_unless(request, "warehouses.manage")
        if blocked:
            return blocked
        form = WarehouseForm(request.POST)
        if form.is_valid():
            item = form.save()
            messages.success(request, f"تمت إضافة المستودع {item.name}")
            return redirect("warehouses")
    page_obj, export_query = paginate_qs(request, qs)
    return render(
        request,
        "setup/warehouses.html",
        {
            "items": page_obj,
            "page_obj": page_obj,
            "form": form,
            "query": q,
            "export_query": export_query,
        },
    )


@login_required
@require_perm("warehouses.view")
def export_warehouses(request):
    q = request.GET.get("q") or ""
    qs = Warehouse.objects.all().order_by("code")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    rows = [[i, item.code, item.name] for i, item in enumerate(qs, 1)]
    return simple_excel(
        "المستودعات",
        ["م", "رقم المستودع", "اسم المستودع"],
        rows,
        [6, 18, 28],
        f"warehouses-{timezone.localdate().isoformat()}.xlsx",
    )


@login_required
@require_perm("warehouses.manage")
def warehouse_edit(request, pk):
    request.active_nav = "setup"
    request.setup_tab = "warehouses"
    item = get_object_or_404(Warehouse, pk=pk)
    form = WarehouseForm(instance=item)
    if request.method == "POST":
        form = WarehouseForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f"تم تعديل المستودع {item.name}")
            return redirect("warehouses")
    return render(request, "setup/warehouse_edit.html", {"form": form, "item": item})


@login_required
@require_POST
@require_perm("warehouses.manage")
def warehouse_delete(request, pk):
    item = get_object_or_404(Warehouse, pk=pk)
    name = item.name
    try:
        item.delete()
    except ProtectedError:
        messages.warning(request, f"لا يمكن حذف {name} لارتباطه ببيانات أخرى")
        return redirect("warehouses")
    messages.success(request, f"تم حذف المستودع {name}")
    return redirect("warehouses")


@login_required
@require_perm("suppliers.view")
def supplier_list(request):
    request.active_nav = "setup"
    request.setup_tab = "suppliers"
    q = request.GET.get("q") or ""
    qs = Supplier.objects.all()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    form = SupplierForm()
    if request.method == "POST":
        blocked = deny_unless(request, "suppliers.manage")
        if blocked:
            return blocked
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f"تمت إضافة المورد {supplier.name}")
            return redirect("suppliers")
    page_obj, export_query = paginate_qs(request, qs)
    return render(
        request,
        "suppliers/list.html",
        {
            "suppliers": page_obj,
            "page_obj": page_obj,
            "form": form,
            "query": q,
            "export_query": export_query,
        },
    )


@login_required
@require_perm("suppliers.view")
def export_suppliers(request):
    q = request.GET.get("q") or ""
    qs = Supplier.objects.all().order_by("code")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "الموردون"
    sheet.sheet_view.rightToLeft = True
    headers = ["م", "رقم المورد", "اسم المورد"]
    header_font = Font(name="Calibri", bold=True, color="1E40AF", size=11)
    header_fill = PatternFill("solid", fgColor="E8F1FF")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="right", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    zebra = PatternFill("solid", fgColor="F8FAFC")
    sheet.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = sheet.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    for index, supplier in enumerate(qs, 1):
        sheet.append([index, supplier.code, supplier.name])
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(index + 1, col)
            cell.alignment = cell_align
            cell.border = thin
            cell.font = Font(name="Calibri", size=11)
            if index % 2 == 0:
                cell.fill = zebra

    for col, width in enumerate([6, 20, 36], 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 28

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="suppliers-{timezone.localdate().isoformat()}.xlsx"'
    )
    return response


@login_required
@require_perm("suppliers.manage")
def supplier_edit(request, pk):
    request.active_nav = "setup"
    request.setup_tab = "suppliers"
    supplier = get_object_or_404(Supplier, pk=pk)
    form = SupplierForm(instance=supplier)
    if request.method == "POST":
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, f"تم تعديل المورد {supplier.name}")
            return redirect("suppliers")
    return render(request, "suppliers/edit.html", {"form": form, "supplier": supplier})


@login_required
@require_POST
@require_perm("suppliers.manage")
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    name = supplier.name
    try:
        supplier.delete()
    except ProtectedError:
        messages.warning(request, f"لا يمكن حذف {name} لارتباطه بطلبات توريد")
        return redirect("suppliers")
    messages.success(request, f"تم حذف المورد {name}")
    return redirect("suppliers")


@login_required
@require_perm("users.manage")
def user_list(request):
    request.active_nav = "users"
    q = request.GET.get("q") or ""
    qs = User.objects.select_related("profile").order_by("username")
    if q:
        qs = qs.filter(
            Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
        )
    form = PortalUserForm()
    if request.method == "POST":
        form = PortalUserForm(request.POST)
        if form.is_valid():
            user = User(username=form.cleaned_data["username"], email=form.cleaned_data.get("email") or "")
            form.apply_name(user)
            user.set_password(form.cleaned_data["password"])
            apply_role_flags(user, form.cleaned_data["role"])
            user.save()
            UserProfile.objects.update_or_create(
                user=user,
                defaults={
                    "role": form.cleaned_data["role"],
                    "phone": form.cleaned_data.get("phone") or "",
                    "branch": form.cleaned_data.get("branch") or "",
                },
            )
            messages.success(request, f"تمت إضافة المستخدم {user.username}")
            return redirect("users")
    page_obj, export_query = paginate_qs(request, qs)
    return render(
        request,
        "users/list.html",
        {
            "users": page_obj,
            "page_obj": page_obj,
            "form": form,
            "query": q,
            "export_query": export_query,
            "role_guide": [(ROLE_LABELS[code], text) for code, text in ROLE_GUIDE],
        },
    )


@login_required
@require_perm("users.manage")
def export_users(request):
    q = request.GET.get("q") or ""
    qs = User.objects.select_related("profile").order_by("username")
    if q:
        qs = qs.filter(
            Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "المستخدمون"
    sheet.sheet_view.rightToLeft = True
    headers = ["م", "رقم المستخدم", "الاسم", "الدور", "البريد", "الهاتف", "الفرع"]
    header_font = Font(name="Calibri", bold=True, color="1E40AF", size=11)
    header_fill = PatternFill("solid", fgColor="E8F1FF")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="right", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    zebra = PatternFill("solid", fgColor="F8FAFC")
    sheet.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = sheet.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    for index, user in enumerate(qs, 1):
        profile = None
        try:
            profile = user.profile
        except UserProfile.DoesNotExist:
            profile = None
        sheet.append(
            [
                index,
                user.username,
                user.get_full_name() or user.username,
                role_label(user) or (getattr(profile, "role", "") or ""),
                user.email,
                getattr(profile, "phone", "") or "",
                getattr(profile, "branch", "") or "",
            ]
        )
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(index + 1, col)
            cell.alignment = cell_align
            cell.border = thin
            cell.font = Font(name="Calibri", size=11)
            if index % 2 == 0:
                cell.fill = zebra

    for col, width in enumerate([6, 22, 22, 18, 32, 18, 22], 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 28

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="users-{timezone.localdate().isoformat()}.xlsx"'
    )
    return response


def _save_user_profile(user, form):
    UserProfile.objects.update_or_create(
        user=user,
        defaults={
            "role": form.cleaned_data["role"],
            "phone": form.cleaned_data.get("phone") or "",
            "branch": form.cleaned_data.get("branch") or "",
        },
    )


@login_required
@require_perm("users.manage")
def user_edit(request, pk):
    request.active_nav = "users"
    user_obj = get_object_or_404(User, pk=pk)
    form = PortalUserForm(instance=user_obj)
    if request.method == "POST":
        form = PortalUserForm(request.POST, instance=user_obj)
        if form.is_valid():
            user_obj.username = form.cleaned_data["username"]
            user_obj.email = form.cleaned_data.get("email") or ""
            form.apply_name(user_obj)
            if form.cleaned_data.get("password"):
                user_obj.set_password(form.cleaned_data["password"])
            apply_role_flags(user_obj, form.cleaned_data["role"])
            user_obj.save()
            _save_user_profile(user_obj, form)
            messages.success(request, f"تم تعديل المستخدم {user_obj.username}")
            return redirect("users")
    return render(request, "users/edit.html", {"form": form, "user_obj": user_obj})


@login_required
@require_POST
@require_perm("users.manage")
def user_delete(request, pk):
    user_obj = get_object_or_404(User, pk=pk)
    if user_obj.pk == request.user.pk:
        messages.warning(request, "لا يمكن حذف حسابك الحالي")
        return redirect("users")
    name = user_obj.username
    user_obj.delete()
    messages.success(request, f"تم حذف المستخدم {name}")
    return redirect("users")
